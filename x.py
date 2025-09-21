#!/usr/bin/env python3
# -*- coding: utf-8 -*-

import os, re, glob, hashlib, subprocess, shutil
from io import BytesIO
from datetime import datetime, timedelta, date
from urllib.parse import urlsplit, urlunsplit, parse_qsl, urlencode

import streamlit as st
import pandas as pd
import requests
from dotenv import load_dotenv
from PIL import Image as PILImage
from openpyxl import Workbook
from openpyxl.drawing.image import Image as XLImage
from openpyxl.styles import Alignment, Font
from pandas.errors import EmptyDataError, ParserError

# =========================
# Lokasi output (SATU folder, konsisten)
# =========================
CSV_DIR = "tweets-data"                      
IMG_DIR = os.path.join(CSV_DIR, "images")
LEGACY_DIRS = ["tweets-data"]               

os.makedirs(CSV_DIR, exist_ok=True)

# ====== Kolom umum dari tweet-harvest ======
DATE_COLS  = ["date", "created_at", "time", "timestamp", "published_at"]
TEXT_COLS  = ["text", "full_text", "content", "caption", "body"]
LIKES_COLS = ["likes", "favorite_count", "like_count", "likes_count", "favoriteCount"]
LINK_COLS  = ["url", "link", "tweet_url", "status_url", "permalink"]
ID_COLS    = ["id", "tweetId", "status_id", "conversation_id", "conversationId"]
MEDIA_COLS = ["photos", "media", "media_urls", "images", "image_urls", "media_url", "media_url_https"]
URL_REGEX  = re.compile(r'https?://[^\s,"]+')

# ====== Helper umum ======
def pick_first_col(df: pd.DataFrame, cands):
    for c in cands:
        if c in df.columns:
            return c
    return None

def to_orig_url(u: str) -> str:
    """Naikkan ke resolusi penuh (name=orig / :orig)."""
    try:
        sp = urlsplit(u)
        if sp.netloc.endswith("pbs.twimg.com") and "/media/" in sp.path:
            path = sp.path
            if ":" in path:
                path = path.split(":")[0] + ":orig"
            qs = dict(parse_qsl(sp.query, keep_blank_values=True))
            qs["name"] = "orig"
            return urlunsplit((sp.scheme, sp.netloc, path, urlencode(qs), sp.fragment))
        if "pbs.twimg.com" in u and "name=" in u:
            return re.sub(r"name=[^&]+", "name=orig", u)
    except Exception:
        pass
    return u

def to_thumb_url(u: str) -> str:
    """Turunkan ke thumbnail (kecil) untuk PREVIEW saja."""
    try:
        sp = urlsplit(u)
        if sp.netloc.endswith("pbs.twimg.com") and "/media/" in sp.path:
            path = sp.path.split(":")[0]  # buang :orig kalau ada
            qs = dict(parse_qsl(sp.query, keep_blank_values=True))
            qs["name"] = "small"
            return urlunsplit((sp.scheme, sp.netloc, path, urlencode(qs), sp.fragment))
    except Exception:
        pass
    return u

def find_image_url(df: pd.DataFrame, row: pd.Series):
    # 1) kolom media spesifik
    for c in MEDIA_COLS:
        if c in df.columns:
            val = row.get(c, None)
            if pd.isna(val):
                continue
            for u in URL_REGEX.findall(str(val)):
                if ("twimg.com/media" in u) or u.lower().endswith((".jpg",".jpeg",".png",".webp")):
                    return to_orig_url(u)
    # 2) scan semua kolom string
    for c in df.columns:
        val = row.get(c, None)
        if pd.isna(val):
            continue
        for u in URL_REGEX.findall(str(val)):
            if ("twimg.com/media" in u) or u.lower().endswith((".jpg",".jpeg",".png",".webp")):
                return to_orig_url(u)
    return None

def build_tweet_link(row: pd.Series, link_col: str | None, id_col: str | None) -> str:
    if link_col:
        val = row.get(link_col, None)
        if isinstance(val, str) and val.startswith("http"):
            return val
    if id_col and not pd.isna(row.get(id_col, None)):
        tid = str(row[id_col]).strip()
        if tid:
            return f"https://x.com/i/web/status/{tid}"
    return ""

def keep_only_original(df: pd.DataFrame) -> pd.DataFrame:
    mask = pd.Series(True, index=df.index)
    for c in ["is_retweet", "retweeted"]:
        if c in df.columns:
            mask &= ~df[c].fillna(False)
    for c in ["in_reply_to_status_id", "in_reply_to_tweet_id", "in_reply_to_user_id", "reply_to"]:
        if c in df.columns:
            mask &= df[c].isna() | (df[c] == 0) | (df[c] == "")
    text_col = pick_first_col(df, TEXT_COLS)
    if text_col:
        mask &= ~df[text_col].fillna("").str.startswith("RT @")
    if "referenced_tweets" in df.columns:
        mask &= ~df["referenced_tweets"].astype(str).str.contains("replied_to|retweeted", case=False, na=False)
    return df[mask].reset_index(drop=True)

def has_image_url(df: pd.DataFrame, row: pd.Series) -> bool:
    return bool(find_image_url(df, row))

# ====== Helper I/O CSV aman & konsisten ======
def _peek_file(path: str, nbytes: int = 2048) -> bytes:
    try:
        with open(path, "rb") as f:
            return f.read(nbytes)
    except Exception:
        return b""

def _looks_like_csv(path: str) -> bool:
    try:
        if not os.path.exists(path) or os.path.getsize(path) == 0:
            return False
        head = _peek_file(path)
        if head.strip().startswith(b"<") and b"<html" in head.lower():
            # HTML error page → bukan CSV
            return False
        return True
    except Exception:
        return False

def _find_csv_after_run(username: str, started_ts: float | None = None) -> str | None:
    """
    Cari CSV yang valid (>0 byte & bukan HTML), prefer nama <username>.csv,
    scan di CSV_DIR, LEGACY_DIRS, dan root. Hanya ambil yang dimodifikasi setelah started_ts (kalau diberikan).
    """
    preferred = [os.path.join(CSV_DIR, f"{username}.csv")]
    for d in LEGACY_DIRS:
        preferred.append(os.path.join(d, f"{username}.csv"))

    pools = []
    pools += preferred
    pools += sorted(glob.glob(os.path.join(CSV_DIR, "*.csv")), key=os.path.getmtime, reverse=True)
    for d in LEGACY_DIRS:
        pools += sorted(glob.glob(os.path.join(d, "*.csv")), key=os.path.getmtime, reverse=True)
    pools += sorted(glob.glob("*.csv"), key=os.path.getmtime, reverse=True)

    best = None
    for p in pools:
        if not os.path.exists(p):
            continue
        try:
            if started_ts and os.path.getmtime(p) < started_ts:
                continue
            if not _looks_like_csv(p):
                continue
            if os.path.basename(p).lower() == f"{username}.csv".lower():
                return p
            if best is None:
                best = p
        except Exception:
            continue
    return best

def _read_csv_safely(path: str) -> pd.DataFrame:
    """
    Baca CSV dengan guard:
    - error kalau 0 byte / HTML
    - fallback encoding & autodetect delimiter
    """
    if not os.path.exists(path):
        raise FileNotFoundError(f"CSV tidak ditemukan: {path}")
    size = os.path.getsize(path)
    if size == 0:
        raise EmptyDataError("CSV kosong (0 byte).")
    head = _peek_file(path)
    if head.strip().startswith(b"<") and b"<html" in head.lower():
        raise ValueError("File bukan CSV (terdeteksi HTML). Cek token/hasil scrape.")

    try:
        return pd.read_csv(path)
    except (EmptyDataError, ParserError):
        try:
            return pd.read_csv(path, encoding="utf-8-sig", engine="python")
        except (EmptyDataError, ParserError):
            return pd.read_csv(path, encoding="utf-8-sig", engine="python", sep=None)

# =========================
# Query & filter
# =========================
def build_query(handle: str, start_date_str: str, end_date_str: str,
                only_original: bool, exclude_quote: bool,
                require_media: bool) -> str:
    """Selalu terapkan HARI AKHIR INKLUSIF WIB (auto until+1)."""
    parts = [f"from:{handle}"]
    if only_original:
        parts += ["-filter:replies", "-filter:retweets"]
    if exclude_quote:
        parts += ["-filter:quote"]
    if require_media:
        parts += ["filter:images"]
    if start_date_str:
        parts.append(f"since:{start_date_str}")
    if end_date_str:
        until_plus = (datetime.strptime(end_date_str, "%Y-%m-%d") + timedelta(days=1)).strftime("%Y-%m-%d")
        parts.append(f"until:{until_plus}")
    return " ".join(parts)

def postfilter_wib(df: pd.DataFrame, start_date_str: str, end_date_str: str) -> pd.DataFrame:
    """Filter 00:00–23:59 WIB sesuai rentang terpilih (selalu)."""
    date_col = pick_first_col(df, DATE_COLS)
    if not date_col or not (start_date_str or end_date_str):
        return df
    dt_utc = pd.to_datetime(df[date_col], errors="coerce", utc=True)
    tz = "Asia/Pontianak"  # UTC+7 (WIB)
    start_utc = pd.NaT
    end_utc   = pd.NaT
    if start_date_str:
        start_utc = (pd.Timestamp(start_date_str + " 00:00:00", tz=tz).tz_convert("UTC"))
    if end_date_str:
        end_utc   = (pd.Timestamp(end_date_str + " 23:59:59", tz=tz).tz_convert("UTC"))
    mask = pd.Series(True, index=df.index)
    if start_date_str:
        mask &= dt_utc >= start_utc
    if end_date_str:
        mask &= dt_utc <= end_utc
    return df[mask].reset_index(drop=True)

# =========================
# Tweet-harvest runner
# =========================
def run_tweet_harvest(output_dir_or_file: str, search_query: str, limit: int, token: str):
    """
    Kirim FOLDER ke -o agar kompatibel dengan perilaku umum tweet-harvest.
    (Jika output_dir_or_file berakhiran .csv, tetap didukung.)
    """
    from shutil import which
    npx_path = which("npx") or which("npx.cmd") or which("npx.exe")
    if not npx_path:
        return False, "npx tidak ditemukan. Install Node.js 20+."

    is_file = output_dir_or_file.lower().endswith(".csv")
    # Pastikan direktori ada
    if is_file:
        os.makedirs(os.path.dirname(os.path.abspath(output_dir_or_file)), exist_ok=True)
    else:
        os.makedirs(os.path.abspath(output_dir_or_file), exist_ok=True)

    cmd = [npx_path, "--yes", "tweet-harvest", "-o", output_dir_or_file, "-s", search_query, "-l", str(limit)]
    if token:
        cmd += ["--token", token]
    try:
        res = subprocess.run(cmd, capture_output=True, text=True, check=True)
        logs = (res.stdout or "") + ("\n" + res.stderr if res.stderr else "")
        return True, logs
    except subprocess.CalledProcessError as e:
        logs = (e.stdout or "") + ("\n" + e.stderr if e.stderr else "")
        return False, logs

def _harvest_locate_csv_after_run(username: str) -> str | None:
    """
    Setelah tweet-harvest jalan, cari CSV bernama <username>.csv di tweets_data/ atau root.
    Jika ketemu di root, pindahkan ke tweets_data/.
    """
    target = os.path.join(CSV_DIR, f"{username}.csv")
    if os.path.exists(target) and os.path.getsize(target) > 0:
        return target

    candidates = [
        os.path.join(os.getcwd(), f"{username}.csv"),
    ]
    candidates += sorted(glob.glob(os.path.join(CSV_DIR, "*.csv")), key=os.path.getmtime, reverse=True)
    for d in LEGACY_DIRS:
        candidates += sorted(glob.glob(os.path.join(d, "*.csv")), key=os.path.getmtime, reverse=True)
    candidates += sorted(glob.glob("*.csv"), key=os.path.getmtime, reverse=True)

    found = None
    for p in candidates:
        if os.path.exists(p) and os.path.getsize(p) > 0 and p.lower().endswith(".csv"):
            if os.path.basename(p).lower() == f"{username}.csv".lower():
                found = p
                break
            if not found:
                found = p

    if found:
        try:
            os.makedirs(CSV_DIR, exist_ok=True)
            if os.path.abspath(found) != os.path.abspath(target):
                shutil.move(found, target)
            return target
        except Exception:
            return found

    return None

# =========================
# Ekspor Excel: Tanggal | Gambar | Link | Caption | Like
# =========================
def export_excel_5cols(mini: pd.DataFrame, username: str,
                       keep_full_image_in_excel: bool,
                       save_originals_to_disk: bool,
                       img_max_w_px: int = 320,
                       img_max_h_px: int = 320,
                       timeout_sec: int = 15,
                       on_progress=None) -> BytesIO:
    """on_progress(i, n) dipanggil per baris (1-based)."""
    if save_originals_to_disk:
        os.makedirs(IMG_DIR, exist_ok=True)

    wb = Workbook()
    ws = wb.active
    ws.title = "Tweets"
    headers = ["Tanggal", "Gambar", "Link", "Caption", "Like"]
    ws.append(headers)
    for i, h in enumerate(headers, 1):
        ws.cell(row=1, column=i).font = Font(bold=True)
    for col, w in {"A":18,"B":45,"C":55,"D":80,"E":10}.items():
        ws.column_dimensions[col].width = w
    wrap = Alignment(wrap_text=True, vertical="top")

    total = len(mini) if len(mini) > 0 else 1
    rix = 2
    for idx, r in enumerate(mini.itertuples(index=False), start=1):
        r = r._asdict() if hasattr(r, "_asdict") else r
        ws.cell(row=rix, column=1, value=str(r["Tanggal"])).alignment = wrap
        link_val = r["Link"] or ""
        c3 = ws.cell(row=rix, column=3, value=link_val)
        if link_val.startswith("http"):
            c3.hyperlink = link_val; c3.style = "Hyperlink"
        c3.alignment = wrap
        ws.cell(row=rix, column=4, value=str(r["Caption"])).alignment = wrap
        try: like_num = int(r["Like"])
        except Exception: like_num = r["Like"]
        ws.cell(row=rix, column=5, value=like_num).alignment = wrap

        img_url = r["Gambar"]
        if isinstance(img_url, str) and img_url.startswith("http"):
            try:
                resp = requests.get(img_url, timeout=timeout_sec)
                resp.raise_for_status()
                raw = resp.content

                if save_originals_to_disk:
                    h = hashlib.md5(img_url.encode("utf-8")).hexdigest()
                    ext = os.path.splitext(urlsplit(img_url).path)[1].lower() or ".jpg"
                    if ext not in [".jpg",".jpeg",".png",".webp"]:
                        ext = ".jpg"
                    with open(os.path.join(IMG_DIR, f"{h}{ext}"), "wb") as f:
                        f.write(raw)

                pil = PILImage.open(BytesIO(raw)).convert("RGB")
                if not keep_full_image_in_excel:
                    pil.thumbnail((img_max_w_px, img_max_h_px))
                bio = BytesIO(); pil.save(bio, format="PNG"); bio.seek(0)
                xl = XLImage(bio)
                ws.add_image(xl, f"B{rix}")
                ws.row_dimensions[rix].height = int(max(pil.height, 28) * 0.75)
            except Exception:
                pass

        if on_progress:
            on_progress(idx, total)
        rix += 1

    out = BytesIO()
    wb.save(out)
    out.seek(0)
    return out

# =========================
# Streamlit UI
# =========================
load_dotenv()
st.set_page_config(page_title="X Scraper", layout="wide")
st.title("Scrape Postingan X")

with st.sidebar:
    st.header("Pengaturan")
    username = st.text_input("Username (tanpa @ juga boleh)").strip().lstrip("@")
    limit = st.number_input("Limit tweet", min_value=1, max_value=5000, value=200, step=50)

    # Date picker (WIB, akhir inklusif)
    default_start = date.today() - timedelta(days=1)
    default_end   = date.today()
    start_end = st.date_input(
        "Rentang tanggal (WIB)",
        value=(default_start, default_end),
        help="Pilih tanggal awal & akhir. Hari akhir selalu dihitung penuh (00:00–23:59 WIB)."
    )
    if isinstance(start_end, tuple) and len(start_end) == 2:
        start_date_obj, end_date_obj = start_end
    else:
        start_date_obj, end_date_obj = default_start, default_end
    start_date_str = start_date_obj.strftime("%Y-%m-%d")
    end_date_str   = end_date_obj.strftime("%Y-%m-%d")

    st.subheader("Filter")
    only_original = st.checkbox("Hanya tweet asli", value=True,
                                help="Buang replies & retweets; hanya postingan asli akun.")
    exclude_quote = st.checkbox("Exclude quote", value=True,
                                help="Buang quote tweets (postingan yang mengutip tweet lain).")
    require_media = st.checkbox("Hanya yang ada gambar", value=False,
                                help="Ambil hanya tweet yang mengandung gambar.")

    st.subheader("Token")
    token = st.text_input("auth_token (kosong = pakai ENV AUTH_TOKEN)",
                          type="password",
                          value=os.getenv("AUTH_TOKEN",""),
                          help="Masukkan nilai cookie 'auth_token' dari akun X yang login.")

    st.subheader("Ekspor")
    keep_full_image_in_excel = st.checkbox("Embed full-res di Excel (besar)", value=False,
                                           help="Tanam gambar resolusi asli di Excel. Ukuran file bisa sangat besar.")
    save_originals_to_disk = st.checkbox("Simpan gambar original", value=True,
                                         help="Simpan file gambar original ke folder tweets_data/images.")
    output_name = st.text_input("Nama file Excel", value=f"tweets_{(username or 'username').strip()}.xlsx")

    run_btn = st.button("🚀 Scrape & Proses", type="primary")

# state
if "df" not in st.session_state: st.session_state.df = None
if "logs" not in st.session_state: st.session_state.logs = ""

# action
if run_btn:
    try:
        if not username:
            st.error("Username wajib diisi."); st.stop()
        if not token:
            st.error("auth_token kosong (isi di sini atau lewat ENV AUTH_TOKEN)."); st.stop()

        # siapkan direktori output
        os.makedirs(CSV_DIR, exist_ok=True)

        # ===== Progress global =====
        step_txt = st.empty()
        step_bar = st.progress(0)

        # Step 1: build query
        step_txt.info("Langkah 1/5: Menyusun query…")
        query = build_query(username, start_date_str, end_date_str,
                            only_original, exclude_quote, require_media)
        st.write("**Query:**", query)
        step_bar.progress(15)

        # Step 2: scrape (force dir)
        step_txt.info("Langkah 2/5: Menjalankan tweet-harvest…")
        run_started_at = datetime.now().timestamp()

        with st.status("Scraping…", expanded=False) as status:
            ok, logs = run_tweet_harvest(CSV_DIR, query, int(limit), token)  # kirim FOLDER, bukan file
            st.session_state.logs = logs
            if not ok:
                status.update(label="Gagal scrape", state="error")
                st.error("Scraping gagal. Lihat log di bawah."); st.stop()
            status.update(label="Scrape selesai", state="complete")
        step_bar.progress(45)

        # Step 3: temukan CSV (kunci: konsisten ke tweets_data/, tapi tetap dukung legacy satu kali)
        step_txt.info("Langkah 3/5: Membaca CSV…")

        expected_csv = os.path.join(CSV_DIR, f"{username}.csv")

        csv_path = None
        if os.path.exists(expected_csv) and _looks_like_csv(expected_csv):
            csv_path = expected_csv
        else:
            csv_path = _find_csv_after_run(username, started_ts=run_started_at)
            if csv_path and os.path.abspath(os.path.dirname(csv_path)) != os.path.abspath(CSV_DIR):
                # Migrasi: pindahkan ke folder konsisten tweets_data/
                try:
                    os.makedirs(CSV_DIR, exist_ok=True)
                    target = expected_csv
                    if os.path.abspath(csv_path) != os.path.abspath(target):
                        shutil.move(csv_path, target)
                    csv_path = target
                except Exception:
                    pass

        if not csv_path or not os.path.exists(csv_path):
            st.error("CSV tidak ditemukan. Lihat log di bawah.")
            st.write("**Diagnostik lokasi CSV terbaru:**")
            diag_rows = []
            for g in [os.path.join(CSV_DIR, "*.csv"), "*.csv"] + [os.path.join(d, "*.csv") for d in LEGACY_DIRS]:
                for p in glob.glob(g):
                    try:
                        diag_rows.append({
                            "path": p,
                            "size_bytes": os.path.getsize(p),
                            "modified": datetime.fromtimestamp(os.path.getmtime(p)).strftime("%Y-%m-%d %H:%M:%S"),
                        })
                    except Exception:
                        pass
            if diag_rows:
                st.dataframe(pd.DataFrame(sorted(diag_rows, key=lambda r: r["modified"], reverse=True)))
            else:
                st.info("Tidak ada file CSV terdeteksi di tweets_data/ atau folder kerja.")
            st.stop()

        # Diagnostik ringkas
        st.write("CSV path:", csv_path)
        st.write("CSV size (bytes):", os.path.getsize(csv_path))

        # Baca CSV dengan guard tahan banting
        try:
            df = _read_csv_safely(csv_path)
        except Exception as e:
            st.error(f"Gagal membaca CSV: {e}")
            st.stop()

        step_bar.progress(60)

        # Step 4: filter WIB + original + media
        step_txt.info("Langkah 4/5: Memfilter data (WIB & jenis postingan)…")
        df = postfilter_wib(df, start_date_str, end_date_str)
        if only_original or exclude_quote:
            df = keep_only_original(df)
            if exclude_quote and "referenced_tweets" in df.columns:
                df = df[~df["referenced_tweets"].astype(str).str.contains("quoted", case=False, na=False)].reset_index(drop=True)
        if require_media and len(df):
            df = df[df.apply(lambda r: has_image_url(df, r), axis=1)].reset_index(drop=True)
        step_bar.progress(75)

        # Step 5: bangun tabel 5 kolom
        step_txt.info("Langkah 5/5: Menyusun tabel preview…")
        date_col  = pick_first_col(df, DATE_COLS)
        text_col  = pick_first_col(df, TEXT_COLS)
        likes_col = pick_first_col(df, LIKES_COLS)
        link_col  = pick_first_col(df, LINK_COLS)
        id_col    = pick_first_col(df, ID_COLS)

        rows = []
        n = max(len(df), 1)
        row_bar = st.progress(0)
        for i, (_, r) in enumerate(df.iterrows(), start=1):
            tanggal = r.get(date_col, "") if date_col else ""
            caption = r.get(text_col, "") if text_col else ""
            likes   = r.get(likes_col, 0) if likes_col else 0
            link    = build_tweet_link(r, link_col, id_col)
            imgurl  = find_image_url(df, r)
            rows.append({"Tanggal": tanggal, "Gambar": imgurl, "Link": link, "Caption": caption, "Like": likes})
            if i % max(n // 20, 1) == 0:
                row_bar.progress(min(int(i / n * 100), 100))
        row_bar.progress(100)
        step_bar.progress(100)
        step_txt.success("Selesai menyusun tabel.")
        mini = pd.DataFrame(rows)

        st.session_state.df = mini
        st.success(f"Sukses. Baris setelah filter: {len(mini)}")

    except Exception as e:
        st.exception(e)

# ===== Preview & Download =====
if st.session_state.df is not None and len(st.session_state.df):
    mini = st.session_state.df.copy()

    # pakai thumbnail kecil untuk PREVIEW
    preview = mini.copy()
    preview["Gambar"] = preview["Gambar"].apply(lambda u: to_thumb_url(u) if isinstance(u, str) else u)

    st.subheader("Preview (5 kolom, gambar thumbnail)")

    add_height = len(preview) > 5
    table_height = 420  # px

    supports_imagecol = hasattr(st, "column_config") and hasattr(st.column_config, "ImageColumn")
    if supports_imagecol:
        kwargs = {
            "use_container_width": True,
            "column_config": {
                "Gambar": st.column_config.ImageColumn(
                    "Gambar", help="Thumbnail pratayang. Ekspor Excel tetap pakai full-res."
                ),
                "Link": st.column_config.LinkColumn("Link"),
                "Like": st.column_config.NumberColumn("Like", format="%d"),
            },
        }
        if add_height:
            kwargs["height"] = table_height
        st.dataframe(preview, **kwargs)
    else:
        # Fallback HTML (untuk Streamlit lama)
        import html as ihtml
        html_df = preview.copy()

        def img_tag(u):
            return f'<img src="{ihtml.escape(u)}" style="max-height:64px;max-width:64px" />' \
                   if isinstance(u, str) and u.startswith("http") else ""

        def link_tag(u):
            if isinstance(u, str) and u.startswith("http"):
                safe = ihtml.escape(u)
                return f'<a href="{safe}" target="_blank">{safe}</a>'
            return ihtml.escape(str(u)) if u is not None else ""

        html_df["Gambar"] = html_df["Gambar"].map(img_tag)
        html_df["Link"]   = html_df["Link"].map(link_tag)

        html_table = html_df.to_html(escape=False, index=False)
        if add_height:
            st.markdown(f'<div style="max-height:{table_height}px; overflow-y:auto">{html_table}</div>', unsafe_allow_html=True)
        else:
            st.markdown(html_table, unsafe_allow_html=True)

    st.caption(f"Total baris: {len(preview)}")

    # Download CSV (server file tetap di tweets_data/)
    st.download_button(
        "⬇️ Download CSV",
        data=mini.to_csv(index=False).encode("utf-8"),
        file_name="tweets.csv",
        mime="text/csv"
    )

    with st.spinner("Membuat Excel…"):
        export_bar = st.progress(0)
        def on_prog(i, n): export_bar.progress(min(int(i/n*100), 100))
        excel_bytes = export_excel_5cols(
            mini=mini,
            username=username,
            keep_full_image_in_excel=keep_full_image_in_excel,
            save_originals_to_disk=save_originals_to_disk,
            on_progress=on_prog,
        )
        export_bar.progress(100)

    st.download_button(
        "⬇️ Download Excel",
        data=excel_bytes.getvalue(),
        file_name=(output_name or f"tweets_{username}.xlsx"),
        mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    )

st.divider()
with st.expander("Log npx / debug"):
    st.code(st.session_state.logs or "(tidak ada log)")
