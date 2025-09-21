#!/usr/bin/env python3
# -*- coding: utf-8 -*-

import json, re, csv, io
from datetime import datetime, date
from dateutil import tz
import requests
import instaloader
import streamlit as st
import pandas as pd

HOMEPAGE = "https://www.instagram.com/"

# ================== Utils ==================
def K(prefix: str, name: str) -> str:
    return f"{prefix}{name}"

def ts_to_iso(dt_aware, tz_name="Asia/Jakarta"):
    """Konversi datetime aware (UTC dari IG) ke zona WIB -> ISO8601."""
    try:
        return dt_aware.astimezone(tz.gettz(tz_name)).isoformat()
    except Exception:
        return dt_aware.isoformat()

def load_cookies_any_from_text(json_text: str):
    """Parse cookies JSON menjadi dict {name:value}."""
    data = json.loads(json_text)
    jar = {}
    if isinstance(data, dict) and "cookies" not in data and "cookie" not in data:
        jar = {k: str(v) for k, v in data.items()}
    elif isinstance(data, dict) and "cookie" in data:
        for part in str(data["cookie"]).split(";"):
            part = part.strip()
            if "=" in part:
                k, v = part.split("=", 1)
                jar[k.strip()] = v.strip()
    elif isinstance(data, list):
        for c in data:
            name = c.get("name")
            value = c.get("value")
            if name and value is not None:
                jar[name] = str(value)
    elif isinstance(data, str):
        for part in data.split(";"):
            part = part.strip()
            if "=" in part:
                k, v = part.split("=", 1)
                jar[k.strip()] = v.strip()
    else:
        raise ValueError("Format cookies tidak dikenali.")
    return jar

def mount_cookies_to_instaloader(L, cookies_dict):
    s = L.context._session
    s.cookies.clear()
    for name, value in cookies_dict.items():
        s.cookies.set(name, value, domain=".instagram.com", path="/")

def get_lsd_and_prime_headers(L):
    """Warm-up untuk LSD token & headers penting."""
    s = L.context._session
    ua = "Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/140.0.0.0 Safari/537.36"
    s.headers["User-Agent"] = ua
    try:
        r = s.get(HOMEPAGE, timeout=30)
    except Exception:
        return
    if r.status_code != 200:
        return
    html = r.text
    m = re.search(r'"LSD",\s*\{\s*"token"\s*:\s*"([^"]+)"', html) or re.search(r'name="lsd"\s+value="([^"]+)"', html)
    if m:
        s.headers["X-FB-LSD"] = m.group(1)
    s.headers.setdefault("X-ASBD-ID", "129477")
    s.headers.setdefault("Accept", "application/json")

def whoami(L):
    try:
        return L.test_login()
    except Exception:
        return None

def is_post_pinned_safe(post) -> bool:
    """Deteksi aman apakah post 'pinned' di berbagai versi instaloader."""
    for attr in ("is_pinned", "pinned"):
        try:
            v = getattr(post, attr, None)
            if isinstance(v, bool):
                return v
        except Exception:
            pass
    try:
        node = getattr(post, "_node", None)
        if isinstance(node, dict):
            return bool(node.get("is_pinned") or node.get("pinned"))
    except Exception:
        pass
    return False

# ================== Export Helpers ==================
def rows_to_csv_bytes(rows):
    buf = io.StringIO()
    writer = csv.DictWriter(buf, fieldnames=["tanggal_post", "gambar", "link_post", "caption", "like", "tipe"])
    writer.writeheader()
    writer.writerows(rows)
    return buf.getvalue().encode("utf-8-sig")

def rows_to_excel_with_images(rows, progress_placeholder=None):
    """Bangun file Excel (xlsx) dengan gambar embedded pada kolom terakhir."""
    try:
        from openpyxl import Workbook
        from openpyxl.drawing.image import Image as XLImage
        from PIL import Image as PILImage
    except Exception:
        raise RuntimeError("Untuk ekspor Excel bergambar, install dulu: pip install openpyxl pillow")

    wb = Workbook()
    ws = wb.active
    ws.title = "IG Posts"

    headers = ["tanggal_post", "caption", "like", "link_post", "tipe", "gambar"]
    ws.append(headers)

    ws.column_dimensions["A"].width = 25
    ws.column_dimensions["B"].width = 60
    ws.column_dimensions["C"].width = 10
    ws.column_dimensions["D"].width = 42
    ws.column_dimensions["E"].width = 16
    ws.column_dimensions["F"].width = 25

    n = len(rows)
    pbar = None
    if progress_placeholder:
        pbar = progress_placeholder.progress(0.0, text="📦 Membuat Excel…")

    for idx, r in enumerate(rows, start=2):
        ws.cell(row=idx, column=1, value=r.get("tanggal_post", ""))
        ws.cell(row=idx, column=2, value=r.get("caption", ""))
        ws.cell(row=idx, column=3, value=r.get("like", 0))

        link = r.get("link_post", "")
        c = ws.cell(row=idx, column=4, value=link)
        if link:
            c.hyperlink = link
            c.style = "Hyperlink"

        ws.cell(row=idx, column=5, value=r.get("tipe", ""))

        url = r.get("gambar", "")
        if url:
            try:
                resp = requests.get(url, timeout=30)
                resp.raise_for_status()
                img_bytes = io.BytesIO(resp.content)
                pil_img = PILImage.open(img_bytes)
                pil_img.thumbnail((320, 320))
                out = io.BytesIO()
                pil_img.save(out, format="PNG")
                out.seek(0)

                ws.row_dimensions[idx].height = 180
                xl_img = XLImage(out)
                ws.add_image(xl_img, f"F{idx}")
            except Exception:
                ws.cell(row=idx, column=6, value=url)

        if pbar and (idx % 10 == 0 or idx == n + 1):
            pbar.progress(min(1.0, (idx - 1) / max(1, n)), text=f"📦 Membuat Excel… ({idx-1}/{n})")

    out = io.BytesIO()
    wb.save(out)
    out.seek(0)
    if pbar:
        pbar.progress(1.0, text="✅ Excel siap diunduh")
    return out.getvalue()

# ================== Core Scraper ==================
def scrape_posts_range(
    L,
    target_username: str,
    limit: int | None = 200,
    d1: date | None = None,
    d2: date | None = None,
    album_all: bool = True,
    polite_break_after_non_pinned_older: int | None = 20
):
    profile = instaloader.Profile.from_username(L.context, target_username)
    wib = tz.gettz("Asia/Jakarta")

    def day_start_wib(d: date | None):
        if not d: return None
        return datetime(d.year, d.month, d.day, 0, 0, 0, tzinfo=wib)

    def day_end_wib(d: date | None):
        if not d: return None
        return datetime(d.year, d.month, d.day, 23, 59, 59, tzinfo=wib)

    # Normalisasi rentang tanggal
    lower_day = upper_day = None
    if d1 and d2:
        lower_day, upper_day = (min(d1, d2), max(d1, d2))
    elif d1 or d2:
        lower_day = d1 or d2
        upper_day = d1 or d2

    lower_start = day_start_wib(lower_day) if lower_day else None
    upper_end   = day_end_wib(upper_day)   if upper_day else None

    rows, kept = [], 0
    posts = profile.get_posts()
    # progress placeholder dikelola dari luar (Streamlit), kembalikan rows saja
    for i, post in enumerate(posts, start=1):  # newest → oldest (pinned bisa nongol di atas)
        dt_utc = getattr(post, "date_utc", None) or getattr(post, "date", None)
        if dt_utc is None:
            continue
        if dt_utc.tzinfo is None:
            dt_utc = dt_utc.replace(tzinfo=tz.UTC)

        dt_wib = dt_utc.astimezone(wib)
        pinned = is_post_pinned_safe(post)

        # --- Filter tanggal: SELALU continue, TIDAK PERNAH break (kecuali limit) ---
        if upper_end and dt_wib > upper_end:
            continue
        if lower_start and dt_wib < lower_start:
            # optimasi sopan: hentikan setelah cukup banyak non-pinned yang lebih tua
            if not pinned and polite_break_after_non_pinned_older:
                polite_break_after_non_pinned_older -= 1
                if polite_break_after_non_pinned_older <= 0:
                    break
            continue

        caption = (post.caption or "").replace("\r", " ").replace("\n", " ").strip()
        link_post = f"https://www.instagram.com/p/{post.shortcode}/"
        likes = getattr(post, "likes", 0)

        if getattr(post, "typename", "") == "GraphSidecar":
            try:
                sidecars_iter = post.get_sidecar_nodes()
                if album_all:
                    for idx, node in enumerate(sidecars_iter, start=1):
                        rows.append({
                            "tanggal_post": ts_to_iso(dt_utc),
                            "gambar": getattr(node, "display_url", "") or "",
                            "link_post": link_post,
                            "caption": caption,
                            "like": likes,
                            "tipe": f"album_gambar_{idx}",
                        })
                        kept += 1
                        if (limit is not None) and (kept >= limit):
                            return rows
                else:
                    first_node = next(sidecars_iter, None)
                    if first_node is not None:
                        rows.append({
                            "tanggal_post": ts_to_iso(dt_utc),
                            "gambar": getattr(first_node, "display_url", "") or "",
                            "link_post": link_post,
                            "caption": caption,
                            "like": likes,
                            "tipe": "album_pertama",
                        })
                        kept += 1
            except Exception:
                rows.append({
                    "tanggal_post": ts_to_iso(dt_utc),
                    "gambar": getattr(post, "url", "") or "",
                    "link_post": link_post,
                    "caption": caption,
                    "like": likes,
                    "tipe": "album_fallback",
                })
                kept += 1

        elif getattr(post, "is_video", False):
            rows.append({
                "tanggal_post": ts_to_iso(dt_utc),
                "gambar": getattr(post, "url", "") or "",  # cover video
                "link_post": link_post,
                "caption": caption,
                "like": likes,
                "tipe": "video",
            })
            kept += 1

        else:
            rows.append({
                "tanggal_post": ts_to_iso(dt_utc),
                "gambar": getattr(post, "url", "") or "",
                "link_post": link_post,
                "caption": caption,
                "like": likes,
                "tipe": "foto",
            })
            kept += 1

        if (limit is not None) and (kept >= limit):
            break

    return rows

# ================== Streamlit UI (dibungkus) ==================
def render_app(key_prefix: str = "ig_"):
    # Hindari error duplikat set_page_config saat dipanggil dari hub
    try:
        st.set_page_config(page_title="Instagram Scraper (Streamlit + Instaloader)", layout="wide")
    except Exception:
        pass

    st.title("📸 Instagram Scraper — Streamlit + Instaloader")

    # --- Session State defaults (pakai prefix) ---
    rows_key = K(key_prefix, "rows")
    df_key   = K(key_prefix, "df")
    last_user_key = K(key_prefix, "last_username")
    if rows_key not in st.session_state: st.session_state[rows_key] = []
    if df_key   not in st.session_state: st.session_state[df_key] = pd.DataFrame(columns=["tanggal_post","gambar","link_post","caption","like","tipe"])
    if last_user_key not in st.session_state: st.session_state[last_user_key] = ""

    with st.expander("1) Upload / Input Cookies JSON", expanded=True):
        cookies_file = st.file_uploader("Upload cookies JSON", type=["json"], key=K(key_prefix, "cookies_uploader"))
        cookies_text_input = st.text_area(
            "Atau paste JSON cookies di sini (kalau file juga diupload, file diprioritaskan)",
            height=150,
            key=K(key_prefix, "cookies_text")
        )
        st.caption("Minimal **sessionid** & **csrftoken**. Export penuh via Cookie-Editor lebih stabil.")

    use_date_filter = st.checkbox("Enable date filter (inclusive WIB range)", value=False, key=K(key_prefix, "chk_date"))

    c1, c2, c3, c4 = st.columns([2, 2, 2, 2])
    with c1:
        username = st.text_input(
            "Username target (tanpa @)",
            value=st.session_state[last_user_key] or "bpskabupatenpasuruan",
            key=K(key_prefix, "inp_user"))
    with c2:
        limit = st.number_input(
            "Maks. posting diambil",
            min_value=1, max_value=100000, value=100, step=10,
            disabled=use_date_filter, key=K(key_prefix, "inp_limit"))
    with c3:
        start_date = st.date_input("Start date (optional)", value=None, disabled=not use_date_filter, key=K(key_prefix, "inp_start"))
    with c4:
        end_date = st.date_input("End date (optional)", value=None, disabled=not use_date_filter, key=K(key_prefix, "inp_end"))

    album_all = st.checkbox("Ambil semua gambar dari album (carousel)?", value=True, key=K(key_prefix, "chk_album"))

    run = st.button("🚀 Jalankan Scrape", type="primary", use_container_width=True, key=K(key_prefix, "btn_run"))

    status_ph = st.empty()
    table_ph = st.empty()
    dl_col1, dl_col2 = st.columns(2)
    gallery_ph = st.container()
    excel_progress = st.empty()

    # --- Action: Scrape
    if run:
        if not cookies_file and not cookies_text_input.strip():
            st.error("Upload atau paste **cookies JSON** terlebih dahulu.")
            st.stop()
        if not username.strip():
            st.error("Isi **username target** terlebih dahulu.")
            st.stop()

        start_dt = start_date if (use_date_filter and isinstance(start_date, date)) else None
        end_dt   = end_date   if (use_date_filter and isinstance(end_date, date))   else None
        effective_limit = None if use_date_filter else int(limit)

        try:
            if cookies_file:
                cookies_text = cookies_file.read().decode("utf-8")
            else:
                cookies_text = cookies_text_input.strip()
            cookies = load_cookies_any_from_text(cookies_text)
        except Exception as e:
            st.error(f"Cookies JSON tidak valid: {e}")
            st.stop()

        with st.spinner("Menyiapkan sesi & login..."):
            L = instaloader.Instaloader(
                download_pictures=False,
                download_videos=False,
                save_metadata=False,
                compress_json=False,
                post_metadata_txt_pattern=None,
                max_connection_attempts=3,
                request_timeout=30,
            )
            mount_cookies_to_instaloader(L, cookies)
            get_lsd_and_prime_headers(L)
            me = whoami(L)
            if me:
                status_ph.success(f"✅ Login via cookies sebagai **@{me}**")
            else:
                status_ph.warning("⚠️ Cookies terpasang tapi tidak terdeteksi login aktif.")

        try:
            rows = scrape_posts_range(
                L,
                target_username=username.strip(),
                limit=effective_limit,
                d1=start_dt,
                d2=end_dt,
                album_all=album_all
            )
        except instaloader.exceptions.QueryReturnedNotFoundException:
            st.error(f"Profil **@{username}** tidak ditemukan / private.")
            rows = []
        except instaloader.exceptions.ConnectionException as e:
            st.error(f"Error koneksi / 403: {e}")
            st.info("Gunakan cookies penuh dan coba lagi beberapa menit.")
            rows = []
        except Exception as e:
            st.error(f"Error tidak terduga: {repr(e)}")
            rows = []

        st.session_state[rows_key] = rows
        st.session_state[df_key] = pd.DataFrame(rows, columns=["tanggal_post", "gambar", "link_post", "caption", "like", "tipe"])
        st.session_state[last_user_key] = username.strip()

    # --- Selalu render dari session_state
    df = st.session_state[df_key]
    rows = st.session_state[rows_key]
    username_for_file = st.session_state[last_user_key] or "hasil_scrape"

    # Tabel dengan preview gambar
    if df.empty:
        st.info("ℹ️ Belum ada data. Jalankan scrape terlebih dahulu.")
    else:
        try:
            table_ph.dataframe(
                df,
                use_container_width=True,
                hide_index=True,
                key=K(key_prefix, "table"),
                column_config={
                    "tanggal_post": st.column_config.TextColumn("Tanggal Post (WIB)"),
                    "gambar": st.column_config.ImageColumn("Gambar", width="small"),
                    "link_post": st.column_config.LinkColumn("Link Post"),
                    "caption": st.column_config.TextColumn("Caption"),
                    "like": st.column_config.NumberColumn("Like", format="%d"),
                    "tipe": st.column_config.TextColumn("Tipe"),
                },
            )
        except Exception:
            st.write(df)

    # Tombol unduhan
    with dl_col1:
        csv_bytes = rows_to_csv_bytes(rows)
        st.download_button(
            label="⬇️ Download CSV",
            data=csv_bytes,
            file_name=f"{username_for_file}_posts.csv",
            mime="text/csv",
            use_container_width=True,
            key=K(key_prefix, "btn_csv")
        )

    with dl_col2:
        if st.button("⬇️ Build Excel (dengan gambar)", use_container_width=True, disabled=df.empty, key=K(key_prefix, "btn_build_xlsx")):
            try:
                xlsx_bytes = rows_to_excel_with_images(df.to_dict("records"), progress_placeholder=excel_progress)
                st.download_button(
                    label="Klik untuk unduh Excel",
                    data=xlsx_bytes,
                    file_name=f"{username_for_file}_posts.xlsx",
                    mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                    use_container_width=True,
                    key=K(key_prefix, "btn_download_xlsx")
                )
            except Exception as e:
                st.error(str(e))

    # Galeri
    with gallery_ph:
        st.markdown("#### Preview Gambar")
        if rows:
            cols = st.columns(3)
            for i, row in enumerate(rows[:12]):
                with cols[i % 3]:
                    st.image(row["gambar"], caption=f"{row['tanggal_post']} ({row['tipe']})", use_container_width=True)
                    st.markdown(f"[Buka Post]({row['link_post']})")
        else:
            st.caption("Tidak ada gambar untuk ditampilkan.")

# ========== Standalone runner ==========
if __name__ == "__main__":
    render_app(key_prefix="ig_")
